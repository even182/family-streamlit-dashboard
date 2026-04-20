import streamlit as st
import pandas as pd
import requests
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
import numpy as np
import datetime
import re
from openpyxl import load_workbook

st.set_page_config(page_title="Family Portfolio Dashboard", layout="wide")


def safe_secret(key: str, default=""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


# =========================
# OneDrive Excel 同步（與上傳並存）
# =========================
def ensure_excel_from_onedrive(xlsx_path: Path) -> bool:
    url = safe_secret("ONEDRIVE_XLSX_URL", "")
    if not isinstance(url, str) or not url.strip():
        return False
    url = url.strip()

    def add_download_param(u: str) -> str:
        if "download=1" in u:
            return u
        return u + ("&" if "?" in u else "?") + "download=1"

    candidates = [url, add_download_param(url)]
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    last_err = None
    for u in candidates:
        try:
            r = requests.get(
                u,
                timeout=45,
                allow_redirects=True,
                headers={"User-Agent": "Mozilla/5.0"},
            )
            r.raise_for_status()
            content = r.content or b""
            if not content.startswith(b"PK"):
                last_err = RuntimeError(f"下載內容不是 Excel（前 20 bytes={content[:20]!r}）")
                continue
            xlsx_path.write_bytes(content)
            return True
        except Exception as e:
            last_err = e
            continue

    if last_err:
        st.warning(f"OneDrive 下載失敗，將使用既有檔案：{last_err}")
    return False


# =========================
# Google Drive / Google Sheets Excel 同步（與上傳並存）
# =========================
def _to_gdrive_xlsx_download_url(u: str) -> str | None:
    if not isinstance(u, str):
        return None
    u = u.strip()
    if not u:
        return None

    m = re.search(r"/spreadsheets/d/([^/]+)/", u)
    if m:
        sid = m.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"

    m = re.search(r"/file/d/([^/]+)/", u)
    if m:
        fid = m.group(1)
        return f"https://drive.google.com/uc?export=download&id={fid}"

    if "drive.google.com/uc" in u and "id=" in u:
        return u

    return None


def ensure_excel_from_gdrive(xlsx_path: Path) -> bool:
    raw = safe_secret("GOOGLE_SHEETS_URL", "") or safe_secret("GDRIVE_FILE_URL", "")
    if not isinstance(raw, str) or not raw.strip():
        return False

    url = _to_gdrive_xlsx_download_url(raw)
    if not url:
        st.warning("Google Drive 連結格式無法辨識，請確認是 Google Sheets 或 Drive 檔案分享連結。")
        return False

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        r = requests.get(url, timeout=45, allow_redirects=True, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        content = r.content or b""
        if not content.startswith(b"PK"):
            raise RuntimeError(f"下載內容不是 Excel（前 20 bytes={content[:20]!r}）")

        xlsx_path.write_bytes(content)
        return True
    except Exception as e:
        st.warning(f"Google Drive 下載失敗，將使用既有檔案：{e}")
        return False


def _touch_reload_flag(source: str):
    st.session_state["_reload_source"] = source


DATA_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = DATA_DIR / "family_data.xlsx"


def to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(
        s.astype(str)
         .str.replace(",", "", regex=False)
         .str.replace(" ", "", regex=False)
         .replace({"nan": None, "": None}),
        errors="coerce"
    ).fillna(0.0)


def _clean_text(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    return str(x).strip()


def extract_allocation_from_analysis_sheet(xlsx_path: Path, sheet_name: str = "Family"):
    """
    直接從 Excel 工作表讀『分析』區塊，避免 DataFrame 因合併儲存格/空白欄位而錯位。
    以 Excel 原始座標找：分析 / 分類 / 參考現值，因此能正確抓到台幣活儲、台股、台股 ETF、美股、美金儲蓄等列。
    """
    if not xlsx_path.exists():
        return None

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return None

    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]

    def clean(v):
        if v is None:
            return ""
        return str(v).strip()

    anchor = None
    for row in ws.iter_rows():
        for cell in row:
            if clean(cell.value) == "分析":
                anchor = (cell.row, cell.column)
                break
        if anchor:
            break
    if not anchor:
        return None

    ar, ac = anchor

    def find_near(token, r0, r1, c0=None, c1=None):
        r0 = max(1, r0)
        r1 = min(ws.max_row, r1)
        c0 = 1 if c0 is None else max(1, c0)
        c1 = ws.max_column if c1 is None else min(ws.max_column, c1)
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                if clean(ws.cell(r, c).value) == token:
                    return r, c
        return None, None

    cat_r, cat_c = find_near("分類", ar - 3, ar + 12)
    if cat_r is None:
        return None

    val_r, val_c = find_near("參考現值", ar - 3, ar + 12)
    if val_r is None:
        val_r, val_c = find_near("成交金額", ar - 3, ar + 12)
    if val_r is None:
        return None

    items = []
    r = cat_r + 1
    while r <= ws.max_row:
        cat = clean(ws.cell(r, cat_c).value)
        if not cat:
            r += 1
            continue
        if cat == "總計":
            break

        raw = ws.cell(r, val_c).value
        val = pd.to_numeric(str(raw).replace(",", "").strip(), errors="coerce")
        if pd.isna(val):
            r += 1
            continue

        items.append({"分類": cat, "金額": float(val)})
        r += 1

    if not items:
        return None

    alloc = pd.DataFrame(items)
    alloc = alloc[alloc["金額"].fillna(0) != 0].copy()
    return alloc if not alloc.empty else None


def make_allocation_pie_from_analysis(xlsx_path: Path):
    alloc = extract_allocation_from_analysis_sheet(xlsx_path, sheet_name="Family")
    if alloc is None or alloc.empty:
        return None

    fig = px.pie(alloc, names="分類", values="金額", title="資產配置（依 Excel『分析』區塊）")
    fig.update_traces(textposition="inside", textinfo="percent+label")
    fig.update_layout(height=520)
    return fig


def _filter_trade_like_rows(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    code_col = "股票代號" if "股票代號" in d.columns else None
    name_col = "股票名稱" if "股票名稱" in d.columns else ("股票" if "股票" in d.columns else None)

    if code_col:
        code = d[code_col].astype(str).str.strip()
        mask = code.notna() & (code != "") & (code.str.lower() != "nan")
        mask &= ~code.isin(["分類", "總計", "分析"])
        d = d[mask]
    elif name_col:
        name = d[name_col].astype(str).str.strip()
        mask = name.notna() & (name != "") & (name.str.lower() != "nan")
        mask &= ~name.isin(["分類", "總計", "分析"])
        d = d[mask]

    return d


@st.cache_data(show_spinner=False)
def load_data(xlsx_path: Path):
    xls = pd.ExcelFile(xlsx_path)
    family_df = pd.read_excel(xls, "Family")
    acct = pd.read_excel(xls, "Family-帳戶紀錄")
    return family_df, acct


def compute_kpi(family_df: pd.DataFrame):
    df = family_df.copy()
    try:
        df = _filter_trade_like_rows(df)
    except Exception:
        pass

    if "分類" in df.columns:
        cat = df["分類"].astype(str).str.strip()
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    invested_col = "成交金額"
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"

    invested = to_num(df[invested_col]) if invested_col in df.columns else pd.Series([0.0])
    realized = to_num(df[realized_col]) if realized_col in df.columns else pd.Series([0.0])
    unrealized = to_num(df[unrealized_col]) if unrealized_col in df.columns else pd.Series([0.0])

    total_invested = float(invested.sum())
    total_realized = float(realized.sum())
    total_unrealized = float(unrealized.sum())
    total_pnl = total_realized + total_unrealized
    ret = (total_pnl / total_invested) if total_invested else 0.0

    return total_invested, total_realized, total_unrealized, total_pnl, ret


def make_rank_chart_by_market(family_df: pd.DataFrame, market: str, top_n: int = 10):
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"
    cat_col = "分類"

    df = _filter_trade_like_rows(family_df)

    name_col = "股票名稱" if "股票名稱" in df.columns else ("股票" if "股票" in df.columns else None)
    if name_col is None:
        return None

    if cat_col in df.columns:
        cat = df[cat_col].astype(str).str.strip()
        if market == "美股":
            df = df[cat == "美股"]
        else:
            df = df[cat.isin(["台股", "台股 ETF"])]
    else:
        return None

    if df.empty:
        return None

    df["已實現損益"] = to_num(df[realized_col]) if realized_col in df.columns else 0.0
    df["未實現損益"] = to_num(df[unrealized_col]) if unrealized_col in df.columns else 0.0
    df["總損益"] = df["已實現損益"] + df["未實現損益"]

    agg = (
        df.groupby(name_col, dropna=True)["總損益"]
          .sum()
          .sort_values(ascending=False)
          .head(top_n)
          .reset_index()
          .rename(columns={name_col: "股票", "總損益": "總損益"})
    )

    if agg.empty:
        return None

    bar_colors = np.where(agg["總損益"] >= 0, "#1f77b4", "#d62728")
    bar_text = agg["總損益"].map(lambda v: f"{v:,.0f}")

    fig = go.Figure()
    fig.add_bar(
        x=agg["總損益"],
        y=agg["股票"],
        orientation="h",
        name="總損益",
        marker_color=bar_colors,
        text=bar_text,
        textposition="outside",
    )

    fig.update_layout(
        title=f"{market} 股票別總損益 Top {top_n}",
        height=520,
        margin=dict(t=70),
        showlegend=False,
    )
    fig.update_xaxes(title="總損益", zeroline=True, zerolinewidth=1, zerolinecolor="gray")
    fig.update_yaxes(title="股票", categoryorder="total ascending")
    return fig


def make_timeseries(acct: pd.DataFrame):
    date_col = "日期" if "日期" in acct.columns else acct.columns[0]
    df0 = acct.copy()
    df0[date_col] = pd.to_datetime(df0[date_col], errors="coerce")
    df0 = df0.dropna(subset=[date_col]).sort_values(date_col)

    principal_candidates = ["台幣本金", "TWD本金", "本金(台幣)"]
    cash_candidates = ["台幣現金水位", "台幣現金", "現金水位", "台幣結餘", "結餘"]

    principal_col = next((c for c in principal_candidates if c in df0.columns), None)
    cash_col = next((c for c in cash_candidates if c in df0.columns), None)

    if principal_col is None and cash_col is None:
        return None

    parts = []
    if cash_col is not None:
        tmp = df0[[date_col, cash_col]].copy()
        tmp["值"] = to_num(tmp[cash_col])
        tmp["項目"] = "台幣現金水位"
        parts.append(tmp[[date_col, "值", "項目"]])

    if principal_col is not None:
        tmp = df0[[date_col, principal_col]].copy()
        tmp["值"] = to_num(tmp[principal_col])
        tmp["項目"] = "台幣本金"
        parts.append(tmp[[date_col, "值", "項目"]])

    df = pd.concat(parts, ignore_index=True)

    if df["項目"].nunique() == 1:
        only = df["項目"].iloc[0]
        fig = px.line(df, x=date_col, y="值", title=f"台幣現金水位圖（來源：帳戶紀錄 / {only}）")
        fig.update_layout(height=450, yaxis_title=only, legend_title_text="")
        return fig

    fig = px.line(df, x=date_col, y="值", color="項目", title="台幣現金水位圖（台幣現金水位 vs 台幣本金）")
    fig.update_layout(height=450, legend_title_text="")
    fig.update_yaxes(title_text="金額")
    return fig


def make_yearly_return_combo(family_df: pd.DataFrame, mode: str = "已實現", attrib: str = "A"):
    realized_col = "已實現損益"
    unrealized_col = "未實現損益"

    buy_date_col = "買進日期" if "買進日期" in family_df.columns else None
    sell_date_col = "賣出日期" if "賣出日期" in family_df.columns else None

    if realized_col not in family_df.columns:
        return None

    df = _filter_trade_like_rows(family_df).copy()

    current_year = datetime.date.today().year
    min_year = 2000
    max_year = current_year

    def _clean_year_series(s: pd.Series) -> pd.Series:
        y = pd.to_numeric(s, errors="coerce")
        return y.where((y >= min_year) & (y <= max_year))

    if buy_date_col is not None:
        df[buy_date_col] = pd.to_datetime(df[buy_date_col], errors="coerce")
    if sell_date_col is not None:
        df[sell_date_col] = pd.to_datetime(df[sell_date_col], errors="coerce")

    sold = df.copy()
    if sell_date_col is not None:
        sold = sold[sold[sell_date_col].notna()].copy()

    yearly_realized = None

    if mode == "已實現":
        if attrib == "A":
            if sell_date_col is None:
                return None
            sold["年度"] = _clean_year_series(sold[sell_date_col].dt.year)
            sold = sold[sold["年度"].notna()].copy()
            sold["年度收益"] = to_num(sold[realized_col])
            yearly_realized = sold.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        elif attrib == "B":
            if buy_date_col is None:
                return None
            sold = sold[sold[buy_date_col].notna()].copy()
            sold["年度"] = _clean_year_series(sold[buy_date_col].dt.year)
            sold = sold[sold["年度"].notna()].copy()
            sold["年度收益"] = to_num(sold[realized_col])
            yearly_realized = sold.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        elif attrib == "C":
            if buy_date_col is None or sell_date_col is None:
                return None

            d = sold[sold[buy_date_col].notna() & sold[sell_date_col].notna()].copy()
            if d.empty:
                return None

            pnl = to_num(d[realized_col]).fillna(0.0).to_numpy()
            rows = []

            for i, r in enumerate(d.itertuples(index=False)):
                b = getattr(r, buy_date_col)
                s = getattr(r, sell_date_col)
                if pd.isna(b) or pd.isna(s):
                    continue

                b = pd.Timestamp(b).normalize()
                s = pd.Timestamp(s).normalize()

                if b.year < min_year:
                    b = pd.Timestamp(f"{min_year}-01-01")
                if s.year > max_year:
                    s = pd.Timestamp(f"{max_year}-12-31")

                if s.year < min_year or b.year > max_year:
                    continue

                if s < b:
                    y = s.year
                    if min_year <= y <= max_year:
                        rows.append((y, float(pnl[i])))
                    continue

                total_days = max((s - b).days + 1, 1)
                start_year = max(b.year, min_year)
                end_year = min(s.year, max_year)

                for y in range(start_year, end_year + 1):
                    seg_start = max(b, pd.Timestamp(f"{y}-01-01"))
                    seg_end = min(s, pd.Timestamp(f"{y}-12-31"))
                    seg_days = (seg_end - seg_start).days + 1
                    if seg_days <= 0:
                        continue
                    rows.append((y, float(pnl[i]) * (seg_days / total_days)))

            if not rows:
                return None

            tmp = pd.DataFrame(rows, columns=["年度", "年度收益"])
            tmp["年度"] = _clean_year_series(tmp["年度"])
            tmp = tmp[tmp["年度"].notna()].copy()
            yearly_realized = tmp.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        else:
            return None

        yearly = yearly_realized

    else:
        if sell_date_col is None:
            return None

        sold2 = df[df[sell_date_col].notna()].copy()
        sold2["年度"] = _clean_year_series(sold2[sell_date_col].dt.year)
        sold2 = sold2[sold2["年度"].notna()].copy()
        sold2["年度收益"] = to_num(sold2[realized_col])
        yearly_realized = sold2.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")

        open_pos = df[df[sell_date_col].isna()].copy()
        if unrealized_col in open_pos.columns and not open_pos.empty:
            open_pos["年度"] = current_year
            open_pos["年度收益"] = to_num(open_pos[unrealized_col])
            yearly_unrealized = open_pos.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
            yearly = pd.concat([yearly_realized, yearly_unrealized], ignore_index=True)
            yearly = yearly.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
        else:
            yearly = yearly_realized

    if yearly is None or yearly.empty:
        return None

    yearly["年度"] = _clean_year_series(yearly["年度"])
    yearly = yearly[yearly["年度"].notna()].copy()
    if yearly.empty:
        return None

    yearly = yearly.groupby("年度", as_index=False)["年度收益"].sum().sort_values("年度")
    yearly["年度"] = yearly["年度"].astype(int)
    yearly["累積收益"] = yearly["年度收益"].cumsum()
    yearly["累積標籤"] = yearly["累積收益"].map(lambda v: f"{v:,.0f}")
    yearly["年度標籤"] = yearly["年度收益"].map(lambda v: f"{v:,.0f}")

    fig = go.Figure()
    bar_colors = np.where(yearly["年度收益"] >= 0, "#1f77b4", "#d62728")
    bar_text_pos = ["outside" if v >= 0 else "inside" for v in yearly["年度收益"]]

    fig.add_bar(
        x=yearly["年度"].astype(str),
        y=yearly["年度收益"],
        name="年度收益",
        marker_color=bar_colors,
        text=yearly["年度標籤"],
        textposition=bar_text_pos,
        yaxis="y",
    )

    fig.add_trace(go.Scatter(
        x=yearly["年度"].astype(str),
        y=yearly["累積收益"],
        name="累積收益",
        mode="lines+markers+text",
        text=yearly["累積標籤"],
        textposition="top center",
        yaxis="y2"
    ))

    left_max = float(max(yearly["年度收益"].max(), 0))
    left_min = float(min(yearly["年度收益"].min(), 0))
    left_pad = max((left_max - left_min) * 0.15, 1000)

    right_max = float(yearly["累積收益"].max())
    right_min = float(yearly["累積收益"].min())
    right_pad = max((right_max - right_min) * 0.15, 1000)

    title_suffix = mode if mode != "已實現" else f"{mode}（{attrib}）"
    fig.update_layout(
        title=f"投資收益（年度 vs 累積）— {title_suffix}",
        xaxis=dict(title="年度", type="category"),
        yaxis=dict(title="年度收益", range=[left_min - left_pad, left_max + left_pad]),
        yaxis2=dict(
            title="累積收益",
            overlaying="y",
            side="right",
            showgrid=False,
            range=[right_min - right_pad, right_max + right_pad],
        ),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
        height=520,
        margin=dict(t=80)
    )
    return fig


view_mode = st.sidebar.radio("顯示內容", ["圖表", "交易明細"], index=0)

source = st.session_state.pop("_reload_source", None)
need_fetch = (not XLSX_PATH.exists()) or (source in ("gdrive", "onedrive"))
if need_fetch:
    fetched = False
    if source == "gdrive":
        fetched = ensure_excel_from_gdrive(XLSX_PATH)
        if (not fetched) and safe_secret("ONEDRIVE_XLSX_URL"):
            fetched = ensure_excel_from_onedrive(XLSX_PATH)
    elif source == "onedrive":
        fetched = ensure_excel_from_onedrive(XLSX_PATH)
        if (not fetched) and (safe_secret("GOOGLE_SHEETS_URL") or safe_secret("GDRIVE_FILE_URL")):
            fetched = ensure_excel_from_gdrive(XLSX_PATH)
    else:
        if safe_secret("GOOGLE_SHEETS_URL") or safe_secret("GDRIVE_FILE_URL"):
            fetched = ensure_excel_from_gdrive(XLSX_PATH)
        if (not fetched) and safe_secret("ONEDRIVE_XLSX_URL"):
            fetched = ensure_excel_from_onedrive(XLSX_PATH)

if not XLSX_PATH.exists():
    st.error("找不到 data/family_data.xlsx。")
    st.stop()

family_df, acct = load_data(XLSX_PATH)

st.title("Family 的投資儀表板")
if XLSX_PATH.exists():
    st.caption(f"資料最後更新時間：{pd.to_datetime(XLSX_PATH.stat().st_mtime, unit='s')}")

total_invested, total_realized, total_unrealized, total_pnl, ret = compute_kpi(family_df)
c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("投入金額", f"{total_invested:,.0f}")
c2.metric("已實現損益", f"{total_realized:,.0f}")
c3.metric("未實現損益", f"{total_unrealized:,.0f}")
c4.metric("總損益", f"{total_pnl:,.0f}")
c5.metric("報酬率", f"{ret*100:,.2f}%")

st.divider()


def render_trade_details(family_df: pd.DataFrame):
    st.subheader("交易明細")

    if "分類" not in family_df.columns:
        st.warning("找不到『分類』欄位，無法依台股/美股切換。")
        st.dataframe(family_df, use_container_width=True)
        return

    market = st.radio("明細篩選", ["台股（含台股 ETF）", "美股", "全部"], horizontal=True)

    df = family_df.copy()
    try:
        df = _filter_trade_like_rows(df)
    except Exception:
        pass

    cat = df["分類"].astype(str).str.strip()
    if market.startswith("台股"):
        df = df[cat.isin(["台股", "台股 ETF"])]
    elif market == "美股":
        df = df[cat == "美股"]
    else:
        df = df[cat.notna() & (cat != "") & (cat.str.lower() != "nan")]

    preferred_cols = [
        "買進日期", "賣出日期", "股票代號", "股票名稱", "分類",
        "股數", "買進價", "賣出價",
        "成交金額", "手續費", "交易稅", "除息",
        "已實現損益", "未實現損益", "參考現值",
        "買進原因", "賣出原因", "備註"
    ]
    cols = [c for c in preferred_cols if c in df.columns]
    df_view = df[cols] if cols else df

    for dc in ["買進日期", "賣出日期"]:
        if dc in df_view.columns:
            df_view[dc] = pd.to_datetime(df_view[dc], errors="coerce")

    st.dataframe(df_view, use_container_width=True, height=560)
    csv = df_view.to_csv(index=False, encoding="utf-8-sig")
    st.download_button("下載明細 CSV", data=csv, file_name="trades.csv", mime="text/csv")


if view_mode == "圖表":
    mode = st.radio("年度收益模式", ["已實現", "含未實現"], horizontal=True)
    attrib = st.radio("年度歸類方式（已實現用）", ["A 賣出年度（實現制）", "B 買進年度（決策歸因）", "C 跨年度攤提（天數分攤）"], horizontal=True)
    attrib_key = attrib.split()[0]

    yearly_fig = make_yearly_return_combo(family_df, mode=mode, attrib=attrib_key)
    if yearly_fig is not None:
        st.plotly_chart(yearly_fig, use_container_width=True)
    else:
        st.info("無法產生『投資收益（年度 vs 累積）』圖表（請確認 Excel 有『賣出日期 / 已實現損益』）。")

    pie = make_allocation_pie_from_analysis(XLSX_PATH)
    if pie is not None:
        st.plotly_chart(pie, use_container_width=True)
    else:
        st.warning("找不到 Excel 內『分析』區塊（含『分類』與『參考現值』）或該區塊資料為空。")

    ts = make_timeseries(acct)
    if ts is not None:
        st.plotly_chart(ts, use_container_width=True)
    else:
        st.warning("帳戶紀錄缺少台幣現金相關欄位（台幣現金/台幣本金/結餘），無法畫台幣現金水位圖。")

    top_market = st.radio("Top10 類型", ["台股（含台股 ETF）", "美股"], horizontal=True)
    if top_market.startswith("台股"):
        top_fig = make_rank_chart_by_market(family_df, market="台股", top_n=10)
        if top_fig is not None:
            st.plotly_chart(top_fig, use_container_width=True)
        else:
            st.info("沒有找到可用的台股資料（Top 10）。")
    else:
        top_fig = make_rank_chart_by_market(family_df, market="美股", top_n=10)
        if top_fig is not None:
            st.plotly_chart(top_fig, use_container_width=True)
        else:
            st.info("沒有找到可用的美股資料（Top 10）。")
else:
    render_trade_details(family_df)
